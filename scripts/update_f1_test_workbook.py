from __future__ import annotations

import argparse
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE = Path.home() / "Downloads" / "F1_Test_Scenarios_Realistic.xlsx"
TARGET = Path(__file__).resolve().parents[1] / "docs" / "F1_Test_Scenarios_Realistic.xlsx"
VERIFIED_ON = date(2026, 8, 15)


def update_case(ws, rows: dict[str, int], test_id: str, **values: str) -> None:
    columns = {
        "case": 7,
        "scenario": 8,
        "precondition": 9,
        "steps": 10,
        "expected": 11,
        "oracle": 12,
        "note": 13,
    }
    row = rows[test_id]
    for key, value in values.items():
        ws.cell(row, columns[key], value)


def build_workbook(source: Path = SOURCE, target: Path = TARGET) -> Path:
    if not source.exists():
        raise FileNotFoundError(f"Không tìm thấy workbook gốc: {source}")

    workbook = openpyxl.load_workbook(source)
    cases = workbook["F1_Test_Cases"]
    rows = {cases.cell(row, 2).value: row for row in range(2, cases.max_row + 1)}

    headers = [
        "Automated test",
        "Test file",
        "Test level",
        "Automation status",
        "Gap / Failure reason",
        "Last verified",
    ]
    template = cases.cell(1, 13)
    for column, header in enumerate(headers, start=14):
        cell = cases.cell(1, column, header)
        cell.font = copy(template.font)
        cell.fill = copy(template.fill)
        cell.border = copy(template.border)
        cell.alignment = copy(template.alignment)
        cell.number_format = template.number_format
        cell.protection = copy(template.protection)

    update_case(
        cases,
        rows,
        "F1-TC-006",
        note="UI đã hiển thị ‘Không tìm thấy địa điểm’; cần bổ sung frontend component test.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-010",
        case="API từ chối request thiếu SOC",
        scenario="Client/API gửi yêu cầu tạo trip nhưng không có initial_soc_percent.",
        precondition="Origin/destination hợp lệ; bỏ field initial_soc_percent khỏi request.",
        steps=("1. POST /api/v1/trips không có initial_soc_percent.\n2. Kiểm tra response và provider call count."),
        expected="API trả VALIDATION_ERROR; không tạo trip và không gọi planning workflow.",
        oracle="HTTP 400; error.code=VALIDATION_ERROR; planner call count=0.",
        note="Frontend luôn có SOC mặc định; đây là API/schema test.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-014",
        expected=("API từ chối có chủ đích; không gọi Goong Directions, station provider hoặc tạo plan."),
        oracle=("HTTP 422; error.code=VALIDATION_ERROR; details.reason=SAME_ORIGIN_DESTINATION."),
        note="Behavior đã được chốt trong contract.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-024",
        expected=(
            "Origin/destination chọn từ Goong có source_type=REAL_API; SOC ban đầu có "
            "source_type=MANUAL; không trình bày các input này như telemetry/OEM live."
        ),
        oracle=("Coordinates selected via Goong Place Detail = REAL_API; initial SOC entered by user = MANUAL."),
    )
    update_case(
        cases,
        rows,
        "F1-TC-027",
        case="Chuyến dài kiểm tra nhu cầu sạc theo dữ liệu live",
        expected=(
            "Planner trả plan an toàn hoặc INFEASIBLE có lý do; nếu có trạm thì route từng "
            "leg dùng Goong waypoint, station CCS2 hợp lệ và mọi arrival SOC ≥15%."
        ),
        oracle="LIVE invariant; không khóa chính xác số trạm hoặc tên trạm.",
        note="Muốn assert đúng một trạm phải dùng fixture version hóa riêng.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-029",
        case="SOC thấp kiểm tra khả năng backtrack gần origin",
        expected=(
            "Nếu có station backtracking được Goong xác minh và vẫn giữ SOC ≥15%, planner có "
            "thể chọn; nếu không có chuỗi an toàn thì trả INFEASIBLE minh bạch."
        ),
        oracle="LIVE invariant; không bắt buộc lần chạy live nào cũng phải có backtracking.",
        note="Fixture F1-TC-030/F1-TC-063 dùng để assert exact hành vi backtracking.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-032",
        expected=(
            "Candidate bị loại dù SOC đủ; nếu không còn phương án an toàn, response chứa "
            "reason code DETOUR_DISTANCE_EXCEEDED."
        ),
        oracle=("Exact rejected candidate; added_distance=10.01 km; reason_codes chứa DETOUR_DISTANCE_EXCEEDED."),
    )
    update_case(
        cases,
        rows,
        "F1-TC-033",
        expected=(
            "Candidate bị loại theo time limit; nếu không còn phương án an toàn, response "
            "chứa reason code DETOUR_TIME_EXCEEDED."
        ),
        oracle=("Exact rejected candidate; added_time=15.1 min; reason_codes chứa DETOUR_TIME_EXCEEDED."),
    )
    update_case(
        cases,
        rows,
        "F1-TC-036",
        case="Hai phương án có chuỗi trạm khác nhau",
        scenario=("Safety engine tìm được hai station chain khác nhau và Goong xác minh route waypoint tương ứng."),
        precondition=("Fixture có 2 station-chain identities khác nhau; mỗi chain có route geometry riêng."),
        expected=(
            "Mỗi option giữ đúng station sequence, route geometry và SOC points của chính nó; map redraw đúng option."
        ),
        oracle="Exact station sequence + geometry hash + SOC points per plan.",
        note="Nhiều geometry cho cùng một station chain nằm ngoài phạm vi F1 hiện tại.",
    )
    update_case(
        cases,
        rows,
        "F1-TC-057",
        precondition=(
            "Fixture đặt station ngay origin hoặc chặng tới station có energy=0; arrival SOC=15.0%; các leg sau hợp lệ."
        ),
        expected=(
            "Boundary đúng 15% được coi là đạt reserve; plan feasible nếu toàn bộ leg còn lại cũng giữ SOC ≥15%."
        ),
        oracle=("Exact: station progress=0 hoặc leg energy=0; arrival SOC=15.0% passes."),
    )
    update_case(
        cases,
        rows,
        "F1-TC-074",
        steps=(
            "1. Chạy cùng fixture 10 lần.\n"
            "2. Loại plan_id, UUID, created_at, retrieved_at và timestamps động.\n"
            "3. So station sequence, route/leg distances, SOC points, consumption và verdict."
        ),
        expected="Kết quả nghiệp vụ sau khi normalize giống nhau trong tolerance đã chốt.",
        oracle=("Exact normalized fixture oracle; không so UUID/timestamp động; không gọi live API."),
    )

    automated = {
        "F1-TC-005": (
            "test_create_trip_returns_ambiguous_location",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-010": (
            "test_create_trip_rejects_missing_soc_at_api_boundary",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-011": (
            "test_create_trip_rejects_invalid_soc",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-014": (
            "test_create_trip_rejects_same_origin_and_destination",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-018": (
            "test_get_current_assumptions_returns_versioned_policy_snapshot",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-019": (
            "test_vf6_profile_exposes_planning_and_official_vehicle_specs",
            "tests/test_api/test_trips.py",
            "API",
        ),
        "F1-TC-026": (
            "test_short_safe_trip_does_not_require_station_provider",
            "tests/test_api/test_planning.py",
            "API integration",
        ),
        "F1-TC-030": (
            "test_station_at_origin_progress_can_be_used_by_low_soc_trip",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
        "F1-TC-032": (
            "test_planner_reports_exact_detour_rejection[distance]",
            "tests/test_core/test_planning_detour.py",
            "Component",
        ),
        "F1-TC-033": (
            "test_planner_reports_exact_detour_rejection[time]",
            "tests/test_core/test_planning_detour.py",
            "Component",
        ),
        "F1-TC-037": (
            "test_incompatible_connector_is_infeasible",
            "tests/test_core/test_feasibility.py",
            "Unit",
        ),
        "F1-TC-040": (
            "test_vinfast_detail_rejects_maintenance_station",
            "tests/test_core/test_vinfast_station_service.py",
            "Unit",
        ),
        "F1-TC-045": (
            "test_charge_target_can_exceed_eighty_when_next_leg_requires_it",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
        "F1-TC-050": (
            "test_dense_short_route_uses_full_adaptive_detail_budget",
            "tests/test_core/test_vinfast_station_service.py",
            "Unit",
        ),
        "F1-TC-056": (
            "test_generate_trip_plan_infeasible_low_soc",
            "tests/test_api/test_planning.py",
            "API integration",
        ),
        "F1-TC-057": (
            "test_station_at_origin_with_soc_exactly_at_reserve_is_feasible",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
        "F1-TC-058": (
            "test_arrival_below_reserve_is_infeasible",
            "tests/test_core/test_feasibility.py",
            "Unit",
        ),
        "F1-TC-059": (
            "test_station_at_origin_with_soc_exactly_at_reserve_is_feasible",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
        "F1-TC-061": (
            "test_no_compatible_corridor_station_is_infeasible_when_charging_is_required",
            "tests/test_core/test_feasibility.py",
            "Unit",
        ),
        "F1-TC-062": (
            "test_unreachable_next_station_is_infeasible",
            "tests/test_core/test_feasibility.py",
            "Unit",
        ),
        "F1-TC-063": (
            "test_backtrack_is_rejected_when_station_arrival_breaks_reserve",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
        "F1-TC-065": (
            "test_routing_failure_returns_provider_error",
            "tests/test_api/test_planning.py",
            "API integration",
        ),
        "F1-TC-069": (
            "test_infeasible_graph_returns_refusal_without_plan_proposal",
            "tests/test_agents/test_graph.py",
            "Component",
        ),
        "F1-TC-074": (
            "test_energy_fixture_is_reproducible_after_excluding_dynamic_identifiers",
            "tests/test_core/test_energy_planning.py",
            "Unit",
        ),
    }

    for row in range(2, cases.max_row + 1):
        test_id = cases.cell(row, 2).value
        data_mode = cases.cell(row, 6).value
        cases.cell(row, 17, "NOT_AUTOMATED")
        cases.cell(
            row,
            18,
            "Chưa có automated test ánh xạ trực tiếp cho toàn bộ oracle của case này.",
        )
        cases.cell(row, 19, VERIFIED_ON)
        if test_id in automated:
            test_name, test_file, level = automated[test_id]
            cases.cell(row, 14, test_name)
            cases.cell(row, 15, test_file)
            cases.cell(row, 16, level)
            if data_mode == "FIXTURE":
                cases.cell(row, 17, "PASS")
                cases.cell(row, 18, None)
            else:
                cases.cell(
                    row,
                    18,
                    "Có test fixture liên quan nhưng LIVE smoke này chưa được tự động hóa end-to-end.",
                )

    gaps = {
        "F1-TC-006": "UI đã sửa; repository chưa có frontend component-test runner.",
        "F1-TC-024": ("Contract/API đã lưu REAL_API và MANUAL; LIVE UI provenance chưa có E2E test."),
        "F1-TC-027": "Có fixture happy-path liên quan; LIVE smoke chưa chạy tự động.",
        "F1-TC-029": "Có unit test backtracking liên quan; LIVE smoke chưa chạy tự động.",
        "F1-TC-036": "Oracle đã đổi sang station-chain identity; chưa có fixture hai chain.",
    }
    for test_id, gap in gaps.items():
        cases.cell(rows[test_id], 18, gap)

    cases.freeze_panes = "A2"
    cases.auto_filter.ref = f"A1:S{cases.max_row}"
    for column, width in {14: 52, 15: 42, 16: 18, 17: 20, 18: 58, 19: 16}.items():
        cases.column_dimensions[get_column_letter(column)].width = width
    for row in cases.iter_rows(min_row=2, max_col=19):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )

    live = workbook["Live_Smoke"]
    live_rows = {live.cell(row, 2).value: row for row in range(2, live.max_row + 1)}
    for test_id in ("F1-TC-024", "F1-TC-027", "F1-TC-029"):
        source_row = rows[test_id]
        target_row = live_rows[test_id]
        live.cell(target_row, 3, cases.cell(source_row, 7).value)
        live.cell(target_row, 6, cases.cell(source_row, 11).value)
        live.cell(target_row, 7, cases.cell(source_row, 12).value)

    rules = workbook["GroundTruth_Rules"]
    rules.cell(
        6,
        2,
        (
            "Goong lỗi → ROUTING_UNAVAILABLE; Open-Meteo lỗi → ENVIRONMENT_DATA_UNAVAILABLE. "
            "VinFast lỗi chỉ trả STATION_DATA_UNAVAILABLE khi hành trình cần sạc; chuyến đi "
            "thẳng đã chứng minh an toàn không phụ thuộc Locator."
        ),
    )
    rules.append(
        (
            "11. Station degradation/cache",
            (
                "Có thể dùng cache locator đã tải khi refresh metadata lỗi và phải giữ "
                "source_updated_at/freshness. Detail lỗi một phần được bỏ qua; nếu không còn "
                "dữ liệu đủ để chứng minh chuỗi sạc an toàn thì fail closed."
            ),
        )
    )
    rules.append(
        (
            "12. Pilot model oracle",
            (
                "Hệ số công suất ×0,85, mục tiêu sạc 80% và các hệ số thời tiết/tải là "
                "oracle của mô hình pilot, không phải ground truth vật lý hoặc cam kết OEM."
            ),
        )
    )
    rules.column_dimensions["A"].width = 34
    rules.column_dimensions["B"].width = 125
    for row in rules.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if "Automation" in workbook.sheetnames:
        del workbook["Automation"]
    automation = workbook.create_sheet("Automation")
    automation.append(("Automation status", "Số case", "Ý nghĩa"))
    automation.append(
        (
            "PASS",
            '=COUNTIF(F1_Test_Cases!$Q$2:$Q$85,"PASS")',
            "Có automated test ánh xạ trực tiếp và đã pass trong lần verify.",
        )
    )
    automation.append(
        (
            "FAIL",
            '=COUNTIF(F1_Test_Cases!$Q$2:$Q$85,"FAIL")',
            "Automated test ánh xạ trực tiếp nhưng đang fail.",
        )
    )
    automation.append(
        (
            "NOT_AUTOMATED",
            '=COUNTIF(F1_Test_Cases!$Q$2:$Q$85,"NOT_AUTOMATED")',
            "Chưa có automated test bao phủ đầy đủ oracle của case.",
        )
    )
    automation.append(())
    automation.append(
        (
            "Quy ước normalize",
            None,
            ("Không so UUID, plan_id, created_at, retrieved_at hoặc timestamp động trong fixture reproducibility."),
        )
    )
    automation.append(
        (
            "LIVE",
            None,
            "Chỉ assert invariant; không khóa tên trạm, số trạm hoặc route geometry theo thời gian.",
        )
    )
    automation.append(("FIXTURE", None, "Dùng oracle exact và provider offline/version hóa."))
    for cell in automation[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    automation.column_dimensions["A"].width = 24
    automation.column_dimensions["B"].width = 14
    automation.column_dimensions["C"].width = 95
    for row in automation.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    automation.freeze_panes = "A2"

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the reviewed F1 acceptance workbook.")
    parser.add_argument("--source", type=Path, default=SOURCE)
    parser.add_argument("--target", type=Path, default=TARGET)
    args = parser.parse_args()
    output = build_workbook(args.source, args.target)
    print(output)
